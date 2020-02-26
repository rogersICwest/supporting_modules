import pandas as pd
import numpy as np
import os
import re
import openpyxl # for modifying xlsx files
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, PatternFill, Font, Border
import xlrd # adding support for older xls files
from IPython.display import HTML, display
import time


# if file is the older version xls, have to convert with this function first
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
def open_xls_as_xlsx(filename, index):  # reading xls into openpyxl workbook
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1

def read_xls_xlsx(filename, sheets_number):
    if filename[-4:] == "xlsx":
        workbook = openpyxl.load_workbook(filename)
        ws = workbook[workbook.sheetnames[sheets_number]]
    elif filename[-4:] == ".xls":
        workbook = open_xls_as_xlsx(filename, sheets_number)
        # because in this case it is converted, it has only the first sheet
        ws = workbook[workbook.sheetnames[0]]
    return ws

def read_val_into_set(filename):
    vals = set()
    try:
        ws = read_xls_xlsx(filename, 0)
    except:
        ws = read_xls_xlsx(filename + "x", 0)

    for i in range(ws.max_row):
        count_space = 0
        for j in range(99):
            val = ws.cell(row=i+1, column=j+1).value
            if val == "" or val is None: count_space += 1
            if count_space >= 5: break
            vals.add(val)
    
    return vals

def progress_bar(value, max=100):
    return HTML("""
        <progress
            value='{value}'
            max='{max}',
            style='width: 100%'
        >
            {value}
        </progress>
    """.format(value=value, max=max))

def get_header(worksheet):
    # assume first row is the header
    # openpyxl is 1-indexed
    header = []
    for i in range(999):
        val = worksheet.cell(row=1, column=i+1).value
        if val == "" or val is None: break
        header.append(val)
    header = [x.lower() for x in header]
    return header

def get_cell_in_partNumber(worksheet, Partnumber_col):
    cells_in_partNumber = []
    for i in range(2,worksheet.max_row+1):
        val = worksheet.cell(row=i, column=Partnumber_col).value
        cells_in_partNumber.append(val)
    return cells_in_partNumber, set(cells_in_partNumber)


def label_rows(main_ws, Partnumber_col):
    # assume that in partnumber column, duplicate part numbers stay together
    # i.e. p1,p1,p1,p2,p2,p3,p3 instead of p1,p1,p2,p1,p2,p3...
    parts = get_cell_in_partNumber(main_ws, Partnumber_col)

    # define color fills
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    greenFill = PatternFill(start_color='FF00FF00',
                          end_color='FF00FF00',
                          fill_type='solid')
    yellowFill = PatternFill(start_color='FFFFFF00',
                          end_color='FFFFFF00',
                          fill_type='solid')
    greyFill = PatternFill(start_color='FF808080',
                          end_color='FFFFFF00',
                          fill_type='solid')
    
    # loop through parts
    N = len(parts[1])
    out = display(progress_bar(0, 100), display_id=True)
    for i, p in enumerate(parts[1]):
        row_ind = parts[0].index(p) + 2
        # debug
        # print("p here is: " + p)
        # print("row_ind here is: " + str(row_ind))
        progress = str(round(100*i/N, 2))
        out.update(progress(i, 100))
        # print("progress: " + progress + "% completed")
        try:
            val_set = read_val_into_set(p + ".xls")
        except:
            # if no file found, label grey
            while main_ws.cell(row=row_ind, column=Partnumber_col).value == p:
                main_ws.cell(row=row_ind, column=Partnumber_col).fill = greyFill
                row_ind += 1
            continue
        count_correct = 0
        # debug
        # print("val set length here is: " + str(len(val_set)))
        # loop through duplicates
        while main_ws.cell(row=row_ind, column=Partnumber_col).value == p:
            # debug
            # print("inside while: " + str(count_correct))
            incorrect = 0
            for i in range(1, Partnumber_col - 5):
                if round(main_ws.cell(row=row_ind, column=i).value,5) not in val_set:
                    # log # of incorrect
                    incorrect += 1
            if incorrect / (Partnumber_col - 5) < 0.2:
                if count_correct == 0:
                    main_ws.cell(row=row_ind, column=Partnumber_col).fill = greenFill
                    count_correct += 1
                else:
                    # TODO: go back and change the other to yellow as well
                    main_ws.cell(row=row_ind, column=Partnumber_col).fill = yellowFill
            else:
                main_ws.cell(row=row_ind, column=Partnumber_col).fill = redFill
            row_ind += 1
